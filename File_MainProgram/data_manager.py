# -*- coding: utf-8 -*-
"""
Module quản lý dữ liệu - Lưu trữ JSON và Xuất Excel
Tác giả: Auto-generated
Mô tả: Quản lý việc lưu/load dữ liệu kết quả phát hiện, tự động dọn dẹp dữ liệu cũ và xuất Excel
"""

import os
import json
from datetime import datetime, timedelta
from threading import Thread, Lock

# Thư viện để xuất Excel với merge cell
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("[CẢNH BÁO] Thư viện openpyxl chưa được cài. Chạy: pip install openpyxl")


class DataManager:
    """
    Lớp quản lý dữ liệu kết quả phát hiện
    - Lưu/Load từ file JSON
    - Tự động xóa dữ liệu cũ
    - Xuất ra Excel với merge cell
    """

    def __init__(self, data_dir=None):
        """
        Khởi tạo DataManager

        Args:
            data_dir: Thư mục lưu dữ liệu. Mặc định là 'data/' trong cùng thư mục với file này
        """
        # Xác định thư mục lưu dữ liệu
        if data_dir is None:
            current_dir = os.path.dirname(os.path.abspath(__file__))
            data_dir = os.path.join(current_dir, 'data')

        self.data_dir = data_dir
        self.json_file = os.path.join(data_dir, 'data_history.json')

        # Lock để đảm bảo thread-safe khi ghi file
        self._lock = Lock()

        # Tạo thư mục nếu chưa tồn tại
        if not os.path.exists(data_dir):
            os.makedirs(data_dir)

        # Load dữ liệu và dọn dẹp dữ liệu cũ
        self.data = self.load_data()
        self.cleanup_old_data()

    def load_data(self):
        """
        Load dữ liệu từ file JSON

        Returns:
            dict: Dictionary với key là ngày (YYYY-MM-DD), value là list các bản ghi
        """
        if os.path.exists(self.json_file):
            try:
                with open(self.json_file, 'r', encoding='utf-8') as f:
                    return json.load(f)
            except (json.JSONDecodeError, IOError) as e:
                print(f"[LỖI] Không thể đọc file JSON: {e}")
                return {}
        return {}

    def _write_json(self):
        """
        Ghi dữ liệu ra file JSON (internal method)
        Được gọi trong thread riêng để không block UI
        """
        with self._lock:
            try:
                with open(self.json_file, 'w', encoding='utf-8') as f:
                    json.dump(self.data, f, ensure_ascii=False, indent=2)
            except IOError as e:
                print(f"[LỖI] Không thể ghi file JSON: {e}")

    def save_record(self, total, passed, failed, result):
        """
        Lưu một bản ghi mới (bất đồng bộ)

        Args:
            total: Tổng số viên phát hiện được
            passed: Số viên đạt (Full)
            failed: Số viên lỗi (Partial, Empty)
            result: Kết quả tổng hợp (OK/NG)

        Returns:
            str: Chuỗi hiển thị cho danh sách UI
        """
        now = datetime.now()
        date_str = now.strftime('%Y-%m-%d')
        time_str = now.strftime('%H:%M:%S')

        # Tạo bản ghi mới
        record = {
            'time': time_str,
            'total': total,
            'passed': passed,
            'failed': failed,
            'result': result
        }

        # Thêm vào dữ liệu
        if date_str not in self.data:
            self.data[date_str] = []
        self.data[date_str].append(record)

        # Ghi file trong thread riêng (async) để không block UI
        Thread(target=self._write_json, daemon=True).start()

        # Trả về chuỗi hiển thị cho UI
        display_str = f"[{time_str}] | Tổng: {total} | Đạt: {passed} | Lỗi: {failed} | Kết quả: {result}"
        return display_str

    def get_today_records(self):
        """
        Lấy danh sách bản ghi của ngày hôm nay

        Returns:
            list: Danh sách các chuỗi hiển thị
        """
        today = datetime.now().strftime('%Y-%m-%d')
        records = self.data.get(today, [])

        display_list = []
        for rec in records:
            display_str = f"[{rec['time']}] | Tổng: {rec['total']} | Đạt: {rec['passed']} | Lỗi: {rec['failed']} | Kết quả: {rec['result']}"
            display_list.append(display_str)

        return display_list

    def cleanup_old_data(self, days=15):
        """
        Xóa dữ liệu cũ hơn số ngày chỉ định

        Args:
            days: Số ngày giữ lại (mặc định 15)
        """
        cutoff_date = datetime.now() - timedelta(days=days)
        cutoff_str = cutoff_date.strftime('%Y-%m-%d')

        # Lọc ra các ngày cần giữ lại
        keys_to_remove = [date for date in self.data.keys() if date < cutoff_str]

        if keys_to_remove:
            for key in keys_to_remove:
                del self.data[key]
            print(f"[THÔNG BÁO] Đã xóa dữ liệu của {len(keys_to_remove)} ngày cũ")
            # Ghi lại file sau khi dọn dẹp
            Thread(target=self._write_json, daemon=True).start()

    def export_to_excel(self, filepath, date_filter=None):
        """
        Xuất dữ liệu ra file Excel với merge cell cho cột Ngày

        Args:
            filepath: Đường dẫn file Excel để lưu
            date_filter: Nếu là chuỗi 'YYYY-MM-DD', chỉ xuất ngày đó. 
                        Nếu là None, xuất toàn bộ dữ liệu.

        Returns:
            bool: True nếu thành công, False nếu thất bại
        """
        if not OPENPYXL_AVAILABLE:
            print("[LỖI] Thư viện openpyxl chưa được cài đặt!")
            return False

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Kết quả phát hiện"

            # Định dạng header
            headers = ['STT', 'Ngày', 'Thời gian', 'Tổng số viên', 'Viên đạt', 'Viên lỗi', 'Kết quả']
            header_font = Font(bold=True, color='FFFFFF')
            header_fill_color = '006666'  # Màu teal
            header_alignment = Alignment(horizontal='center', vertical='center')
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Ghi header
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = thin_border
                # Đặt màu nền cho header
                from openpyxl.styles import PatternFill
                cell.fill = PatternFill(start_color=header_fill_color, end_color=header_fill_color, fill_type='solid')

            # Ghi dữ liệu
            row_num = 2
            stt = 1

            # Biến đếm số lần từng trạng thái kết quả (5 trạng thái)
            count_ok = 0
            count_ng_l = 0
            count_ng_h = 0
            count_missing = 0
            count_wait = 0

            # Lọc các ngày cần xuất
            if date_filter:
                if date_filter in self.data:
                    sorted_dates = [date_filter]
                else:
                    sorted_dates = []
            else:
                # Sắp xếp các ngày theo thứ tự
                sorted_dates = sorted(self.data.keys())

            for date_str in sorted_dates:
                records = self.data[date_str]
                if not records:
                    continue

                # Chuyển đổi định dạng ngày từ YYYY-MM-DD sang DD/MM/YYYY
                date_obj = datetime.strptime(date_str, '%Y-%m-%d')
                display_date = date_obj.strftime('%d/%m/%Y')

                start_row = row_num
                for i, rec in enumerate(records):
                    current_row = row_num
                    ws.cell(row=current_row, column=1, value=stt).alignment = header_alignment
                    ws.cell(row=current_row, column=1).border = thin_border

                    # Chỉ ghi ngày vào hàng đầu tiên của ngày đó
                    if i == 0:
                        ws.cell(row=current_row, column=2, value=display_date).alignment = header_alignment
                        ws.cell(row=current_row, column=2).border = thin_border
                    else:
                        ws.cell(row=current_row, column=2).border = thin_border

                    ws.cell(row=current_row, column=3, value=rec['time']).alignment = header_alignment
                    ws.cell(row=current_row, column=3).border = thin_border

                    ws.cell(row=current_row, column=4, value=rec['total']).alignment = header_alignment
                    ws.cell(row=current_row, column=4).border = thin_border

                    ws.cell(row=current_row, column=5, value=rec['passed']).alignment = header_alignment
                    ws.cell(row=current_row, column=5).border = thin_border

                    ws.cell(row=current_row, column=6, value=rec['failed']).alignment = header_alignment
                    ws.cell(row=current_row, column=6).border = thin_border

                    ws.cell(row=current_row, column=7, value=rec['result']).alignment = header_alignment
                    ws.cell(row=current_row, column=7).border = thin_border

                    # Đếm số lần từng trạng thái kết quả
                    result_val = rec.get('result', '')
                    if result_val == 'OK':
                        count_ok += 1
                    elif result_val == 'NG_L':
                        count_ng_l += 1
                    elif result_val == 'NG_H':
                        count_ng_h += 1
                    elif result_val == 'MISSING':
                        count_missing += 1
                    elif result_val == 'WAIT':
                        count_wait += 1

                    stt += 1
                    row_num += 1

                end_row = row_num - 1

                # Merge cell cột Ngày nếu có nhiều hơn 1 bản ghi trong ngày
                if end_row > start_row:
                    ws.merge_cells(start_row=start_row, start_column=2, end_row=end_row, end_column=2)
                    # Căn trung tâm cho ô đã gộp
                    ws.cell(row=start_row, column=2).alignment = Alignment(horizontal='center', vertical='center')

            # === PHẦN TỔNG HỢP KẾT QUẢ Ở CUỐI ===
            if row_num > 2:  # Chỉ thêm nếu có dữ liệu
                from openpyxl.styles import PatternFill
                
                summary_font = Font(bold=True, size=11)
                summary_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')  # Vàng nhạt
                
                # --- HÀNG 1: Tiêu đề "TỔNG HỢP KẾT QUẢ" + OK + NG_L ---
                row1 = row_num
                ws.merge_cells(start_row=row1, start_column=1, end_row=row1, end_column=3)
                cell_title = ws.cell(row=row1, column=1, value="TỔNG HỢP KẾT QUẢ")
                cell_title.font = Font(bold=True, size=12)
                cell_title.alignment = Alignment(horizontal='center', vertical='center')
                cell_title.border = thin_border
                cell_title.fill = summary_fill

                # OK
                cell_ok = ws.cell(row=row1, column=4, value=f"OK: {count_ok}")
                cell_ok.font = summary_font
                cell_ok.alignment = Alignment(horizontal='center', vertical='center')
                cell_ok.border = thin_border
                cell_ok.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')  # Xanh nhạt

                # NG_L (lỗi nhẹ)
                cell_ng_l = ws.cell(row=row1, column=5, value=f"NG_L: {count_ng_l}")
                cell_ng_l.font = summary_font
                cell_ng_l.alignment = Alignment(horizontal='center', vertical='center')
                cell_ng_l.border = thin_border
                cell_ng_l.fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')  # Vàng nhạt

                # NG_H (lỗi nặng)
                cell_ng_h = ws.cell(row=row1, column=6, value=f"NG_H: {count_ng_h}")
                cell_ng_h.font = summary_font
                cell_ng_h.alignment = Alignment(horizontal='center', vertical='center')
                cell_ng_h.border = thin_border
                cell_ng_h.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')  # Đỏ nhạt

                # MISSING
                cell_miss = ws.cell(row=row1, column=7, value=f"MISSING: {count_missing}")
                cell_miss.font = summary_font
                cell_miss.alignment = Alignment(horizontal='center', vertical='center')
                cell_miss.border = thin_border
                cell_miss.fill = PatternFill(start_color='FFD699', end_color='FFD699', fill_type='solid')  # Cam nhạt

                # --- HÀNG 2: Tổng số lượt kiểm tra ---
                row2 = row_num + 1
                total_count = count_ok + count_ng_l + count_ng_h + count_missing + count_wait
                total_ng = count_ng_l + count_ng_h

                ws.merge_cells(start_row=row2, start_column=1, end_row=row2, end_column=3)
                cell_total_label = ws.cell(row=row2, column=1, value=f"TỔNG LƯỢT: {total_count}")
                cell_total_label.font = Font(bold=True, size=12)
                cell_total_label.alignment = Alignment(horizontal='center', vertical='center')
                cell_total_label.border = thin_border
                cell_total_label.fill = summary_fill

                # Tổng đạt
                cell_dat = ws.cell(row=row2, column=4, value=f"Đạt: {count_ok}")
                cell_dat.font = summary_font
                cell_dat.alignment = Alignment(horizontal='center', vertical='center')
                cell_dat.border = thin_border
                cell_dat.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')

                # Tổng lỗi (NG_L + NG_H)
                cell_loi = ws.cell(row=row2, column=5, value=f"Lỗi: {total_ng}")
                cell_loi.font = summary_font
                cell_loi.alignment = Alignment(horizontal='center', vertical='center')
                cell_loi.border = thin_border
                cell_loi.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

                # Tỷ lệ đạt
                if total_count > 0:
                    ty_le = (count_ok / total_count) * 100
                    ty_le_str = f"Tỷ lệ đạt: {ty_le:.1f}%"
                else:
                    ty_le_str = "Tỷ lệ đạt: --"
                ws.merge_cells(start_row=row2, start_column=6, end_row=row2, end_column=7)
                cell_tyle = ws.cell(row=row2, column=6, value=ty_le_str)
                cell_tyle.font = Font(bold=True, size=11, color='006666')
                cell_tyle.alignment = Alignment(horizontal='center', vertical='center')
                cell_tyle.border = thin_border
                cell_tyle.fill = summary_fill

            # Điều chỉnh độ rộng cột
            column_widths = [6, 15, 12, 15, 12, 12, 15]
            for i, width in enumerate(column_widths, 1):
                ws.column_dimensions[chr(64 + i)].width = width

            # Lưu file
            wb.save(filepath)
            print(f"[THÔNG BÁO] Đã xuất file Excel: {filepath}")
            return True

        except Exception as e:
            print(f"[LỖI] Không thể xuất Excel: {e}")
            return False


# Singleton instance để sử dụng trong toàn ứng dụng
_data_manager_instance = None


def get_data_manager():
    """
    Lấy instance của DataManager (Singleton pattern)

    Returns:
        DataManager: Instance duy nhất của DataManager
    """
    global _data_manager_instance
    if _data_manager_instance is None:
        _data_manager_instance = DataManager()
    return _data_manager_instance
